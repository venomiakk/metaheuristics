import numpy as np
from openpyxl import load_workbook
from algorithm import ParticleSwarmAlgorithm
from plots import plot_ackley, plot_himmelblaus, heatmap, histogram_ackley, histogram_himmelblaus
from datetime import datetime

glb_iterations = 50
glb_no_particles = 50
glb_inertia = 0.5
glb_social = 0.5
glb_cognitive = 0.5
no_runs = 5
excel_file = 'output/res1/res2.xlsx'
plot_folder = 'output/res1'


def format_for_xlsx(title, xs, ys, vs, details):
    data = [[title[0]],
            [title[1]],
            ['', 'x', 'y', 'f(x, y)']]
    for i in range(no_runs):
        new_row = ['', xs[i], ys[i], vs[i]]
        data.append(new_row)
    data.append([])
    data.append(['', 'worst v', 'best v', 'avg v'])
    details.insert(0, '')
    data.append(details)
    return data


def save_to_xlsx(data, file):
    file_path = file
    workbook = load_workbook(file_path)
    sheet = workbook.active

    new_data = data
    new_data.insert(0, [datetime.now()])

    last_row = sheet.max_row

    additional_rows = 3
    start_row = last_row + additional_rows + 1

    for i, row in enumerate(new_data, start=start_row):
        for j, value in enumerate(row, start=1):
            sheet.cell(row=i, column=j, value=value)

    workbook.save(file_path)
    print(f"Saved to: {file_path}")


def experiment1_iters(param, function, xlsx_file=None, plot_file=None):
    xs, ys, vs = [], [], []
    for i in range(no_runs):
        obj = ParticleSwarmAlgorithm(number_of_particles=glb_no_particles, choosen_function=function,
                                     particle_inertia=glb_inertia,
                                     particle_social=glb_social,
                                     particle_cognition=glb_cognitive, stop_condition=0, stop_value=param)
        iters, best_particle = obj.run()
        xs.append(best_particle.x)
        ys.append(best_particle.y)
        vs.append(best_particle.fitness)
        if function == 0:
            f_f = None
            f_h = None
            if plot_file:
                f_f = f'{plot_file}/{i}_ackley_param{param}'
                f_h = f'{plot_file}/{i}_ackley_hist_param{param}'
            plot_ackley(particles=obj.swarm, best_particle=best_particle, iteration=param, inertia=glb_inertia,
                        social=glb_social,
                        cognition=glb_cognitive, no_particles=glb_no_particles, file=f_f)
            histogram_ackley(obj.swarm, file=f_h)
        else:
            f_f = None
            f_h = None
            if plot_file:
                f_f = f'{plot_file}/{i}_himmelblau_param{param}'
                f_h = f'{plot_file}/{i}_himmelblau_hist_param{param}'
            plot_himmelblaus(particles=obj.swarm, best_particle=best_particle, iteration=param,
                             inertia=glb_inertia,
                             social=glb_social,
                             cognition=glb_cognitive, no_particles=glb_no_particles, file=f_f)
            histogram_himmelblaus(obj.swarm, file=f_h)

    worst_v = max(vs)
    best_v = min(vs)
    avg_v = np.average(vs)
    xs = [round(num, 6) for num in xs]
    ys = [round(num, 6) for num in ys]
    vs = [round(num, 6) for num in vs]
    print(f'Results - function: {function}, iters: {param}')
    print(f'x: {xs}')
    print(f'y: {ys}')
    print(f'v: {vs}')
    print(f'best v: {best_v}, avg v: {avg_v}, worst v: {worst_v}')

    func = f'function: {function}'
    param_s = f'iters: {param}'
    title = [func, param_s]
    details = [worst_v, best_v, avg_v]
    data = format_for_xlsx(title, xs, ys, vs, details)
    if xlsx_file:
        save_to_xlsx(data, xlsx_file)


def experiment2_no_particles(param, function, xlsx_file=None, plot_file=None):
    xs, ys, vs = [], [], []
    for i in range(no_runs):
        obj = ParticleSwarmAlgorithm(number_of_particles=param, choosen_function=function, particle_inertia=glb_inertia,
                                     particle_social=glb_social,
                                     particle_cognition=glb_cognitive, stop_condition=0, stop_value=glb_iterations)
        iters, best_particle = obj.run()
        xs.append(best_particle.x)
        ys.append(best_particle.y)
        vs.append(best_particle.fitness)
        if function == 0:
            f_f = None
            f_h = None
            if plot_file:
                f_f = f'{plot_file}/{i}_ackley_param{param}'
                f_h = f'{plot_file}/{i}_ackley_hist_param{param}'
            plot_ackley(particles=obj.swarm, best_particle=best_particle, iteration=glb_iterations, inertia=glb_inertia,
                        social=glb_social,
                        cognition=glb_cognitive, no_particles=param, file=f_f)
            histogram_ackley(obj.swarm, file=f_h)
        else:
            f_f = None
            f_h = None
            if plot_file:
                f_f = f'{plot_file}/{i}_himmelblau_param{param}'
                f_h = f'{plot_file}/{i}_himmelblau_hist_param{param}'
            plot_himmelblaus(particles=obj.swarm, best_particle=best_particle, iteration=glb_iterations,
                             inertia=glb_inertia,
                             social=glb_social,
                             cognition=glb_cognitive, no_particles=param, file=f_f)
            histogram_himmelblaus(obj.swarm, file=f_h)

    worst_v = max(vs)
    best_v = min(vs)
    avg_v = np.average(vs)
    xs = [round(num, 6) for num in xs]
    ys = [round(num, 6) for num in ys]
    vs = [round(num, 6) for num in vs]
    print(f'Results - function: {function}, particles: {param}')
    print(f'x: {xs}')
    print(f'y: {ys}')
    print(f'v: {vs}')
    print(f'best v: {best_v}, avg v: {avg_v}, worst v: {worst_v}')

    func = f'function: {function}'
    param_s = f'particles: {param}'
    title = [func, param_s]
    details = [worst_v, best_v, avg_v]
    data = format_for_xlsx(title, xs, ys, vs, details)
    if xlsx_file:
        save_to_xlsx(data, xlsx_file)


def experiment3_inertia(param, function, xlsx_file=None, plot_file=None):
    xs, ys, vs = [], [], []
    for i in range(no_runs):
        obj = ParticleSwarmAlgorithm(number_of_particles=glb_no_particles, choosen_function=function,
                                     particle_inertia=param,
                                     particle_social=glb_social,
                                     particle_cognition=glb_cognitive, stop_condition=0, stop_value=glb_iterations)
        iters, best_particle = obj.run()
        xs.append(best_particle.x)
        ys.append(best_particle.y)
        vs.append(best_particle.fitness)
        if function == 0:
            f_f = None
            f_h = None
            if plot_file:
                f_f = f'{plot_file}/{i}_ackley_param{param}'
                f_h = f'{plot_file}/{i}_ackley_hist_param{param}'
            plot_ackley(particles=obj.swarm, best_particle=best_particle, iteration=glb_iterations, inertia=param,
                        social=glb_social,
                        cognition=glb_cognitive, no_particles=glb_no_particles, file=f_f)
            histogram_ackley(obj.swarm, file=f_h)
        else:
            f_f = None
            f_h = None
            if plot_file:
                f_f = f'{plot_file}/{i}_himmelblau_param{param}'
                f_h = f'{plot_file}/{i}_himmelblau_hist_param{param}'
            plot_himmelblaus(particles=obj.swarm, best_particle=best_particle, iteration=glb_iterations,
                             inertia=param,
                             social=glb_social,
                             cognition=glb_cognitive, no_particles=glb_no_particles, file=f_f)
            histogram_himmelblaus(obj.swarm, file=f_h)

    worst_v = max(vs)
    best_v = min(vs)
    avg_v = np.average(vs)
    xs = [round(num, 6) for num in xs]
    ys = [round(num, 6) for num in ys]
    vs = [round(num, 6) for num in vs]
    print(f'Results - function: {function}, inertia: {param}')
    print(f'x: {xs}')
    print(f'y: {ys}')
    print(f'v: {vs}')
    print(f'best v: {best_v}, avg v: {avg_v}, worst v: {worst_v}')

    func = f'function: {function}'
    param_s = f'inertia: {param}'
    title = [func, param_s]
    details = [worst_v, best_v, avg_v]
    data = format_for_xlsx(title, xs, ys, vs, details)
    if xlsx_file:
        save_to_xlsx(data, xlsx_file)


def experiment4_cognitive(param, function, xlsx_file=None, plot_file=None):
    xs, ys, vs = [], [], []
    for i in range(no_runs):
        obj = ParticleSwarmAlgorithm(number_of_particles=glb_no_particles, choosen_function=function,
                                     particle_inertia=glb_inertia,
                                     particle_social=glb_social,
                                     particle_cognition=param, stop_condition=0, stop_value=glb_iterations)
        iters, best_particle = obj.run()
        xs.append(best_particle.x)
        ys.append(best_particle.y)
        vs.append(best_particle.fitness)
        if function == 0:
            f_f = None
            f_h = None
            if plot_file:
                f_f = f'{plot_file}/{i}_ackley_param{param}'
                f_h = f'{plot_file}/{i}_ackley_hist_param{param}'
            plot_ackley(particles=obj.swarm, best_particle=best_particle, iteration=glb_iterations, inertia=glb_inertia,
                        social=glb_social,
                        cognition=param, no_particles=glb_no_particles, file=f_f)
            histogram_ackley(obj.swarm, file=f_h)
        else:
            f_f = None
            f_h = None
            if plot_file:
                f_f = f'{plot_file}/{i}_himmelblau_param{param}'
                f_h = f'{plot_file}/{i}_himmelblau_hist_param{param}'
            plot_himmelblaus(particles=obj.swarm, best_particle=best_particle, iteration=glb_iterations,
                             inertia=glb_inertia,
                             social=glb_social,
                             cognition=param, no_particles=glb_no_particles, file=f_f)
            histogram_himmelblaus(obj.swarm, file=f_h)

    worst_v = max(vs)
    best_v = min(vs)
    avg_v = np.average(vs)
    xs = [round(num, 6) for num in xs]
    ys = [round(num, 6) for num in ys]
    vs = [round(num, 6) for num in vs]
    print(f'Results - function: {function}, cognitive: {param}')
    print(f'x: {xs}')
    print(f'y: {ys}')
    print(f'v: {vs}')
    print(f'best v: {best_v}, avg v: {avg_v}, worst v: {worst_v}')

    func = f'function: {function}'
    param_s = f'cognitive: {param}'
    title = [func, param_s]
    details = [worst_v, best_v, avg_v]
    data = format_for_xlsx(title, xs, ys, vs, details)
    if xlsx_file:
        save_to_xlsx(data, xlsx_file)


def experiment5_social(param, function, xlsx_file=None, plot_file=None):
    xs, ys, vs = [], [], []
    for i in range(no_runs):
        obj = ParticleSwarmAlgorithm(number_of_particles=glb_no_particles, choosen_function=function,
                                     particle_inertia=glb_inertia,
                                     particle_social=param,
                                     particle_cognition=glb_cognitive, stop_condition=0, stop_value=glb_iterations)
        iters, best_particle = obj.run()
        xs.append(best_particle.x)
        ys.append(best_particle.y)
        vs.append(best_particle.fitness)
        if function == 0:
            f_f = None
            f_h = None
            if plot_file:
                f_f = f'{plot_file}/{i}_ackley_param{param}'
                f_h = f'{plot_file}/{i}_ackley_hist_param{param}'
            plot_ackley(particles=obj.swarm, best_particle=best_particle, iteration=glb_iterations, inertia=glb_inertia,
                        social=param,
                        cognition=glb_cognitive, no_particles=glb_no_particles, file=f_f)
            histogram_ackley(obj.swarm, file=f_h)
        else:
            f_f = None
            f_h = None
            if plot_file:
                f_f = f'{plot_file}/{i}_himmelblau_param{param}'
                f_h = f'{plot_file}/{i}_himmelblau_hist_param{param}'
            plot_himmelblaus(particles=obj.swarm, best_particle=best_particle, iteration=glb_iterations,
                             inertia=glb_inertia,
                             social=param,
                             cognition=glb_cognitive, no_particles=glb_no_particles, file=f_f)
            histogram_himmelblaus(obj.swarm, file=f_h)

    worst_v = max(vs)
    best_v = min(vs)
    avg_v = np.average(vs)
    xs = [round(num, 6) for num in xs]
    ys = [round(num, 6) for num in ys]
    vs = [round(num, 6) for num in vs]
    print(f'Results - function: {function}, social: {param}')
    print(f'x: {xs}')
    print(f'y: {ys}')
    print(f'v: {vs}')
    print(f'best v: {best_v}, avg v: {avg_v}, worst v: {worst_v}')

    func = f'function: {function}'
    param_s = f'social: {param}'
    title = [func, param_s]
    details = [worst_v, best_v, avg_v]
    data = format_for_xlsx(title, xs, ys, vs, details)
    if xlsx_file:
        save_to_xlsx(data, xlsx_file)


def test_ackley():
    obj = ParticleSwarmAlgorithm(number_of_particles=glb_no_particles, choosen_function=0, particle_inertia=glb_inertia,
                                 particle_social=2,
                                 particle_cognition=0.1, stop_condition=0, stop_value=glb_iterations)
    iters, best_particle = obj.run()
    plot_ackley(particles=obj.swarm, best_particle=best_particle, iteration=iters, inertia=glb_inertia,
                social=glb_social,
                cognition=glb_cognitive, no_particles=glb_no_particles)
    print(best_particle.x, best_particle.y, best_particle.fitness, obj.best_particle_fitness)
    print(obj.best_particle.x, obj.best_particle.y, obj.best_particle.fitness)
    heatmap(obj.all_particles)
    histogram_ackley(obj.swarm)


def test_himmelblaus():
    obj = ParticleSwarmAlgorithm(number_of_particles=glb_no_particles, choosen_function=1, particle_inertia=glb_inertia,
                                 particle_social=2,
                                 particle_cognition=0.1, stop_condition=0, stop_value=glb_iterations)
    iters, best_particle = obj.run()
    plot_himmelblaus(particles=obj.swarm, best_particle=best_particle, iteration=iters, inertia=glb_inertia,
                     social=glb_social,
                     cognition=glb_cognitive, no_particles=glb_no_particles)
    print(best_particle.x, best_particle.y, best_particle.fitness, obj.best_particle_fitness)
    print(obj.best_particle.x, obj.best_particle.y, obj.best_particle.fitness)
    heatmap(obj.all_particles)
    histogram_himmelblaus(obj.swarm)


def run_all_experiments_savedata():
    plt_path = f'{plot_folder}/ex11'
    experiment1_iters(15, 0, xlsx_file=excel_file, plot_file=plt_path)
    experiment1_iters(50, 0, xlsx_file=excel_file, plot_file=plt_path)
    experiment1_iters(100, 0, xlsx_file=excel_file, plot_file=plt_path)

    experiment1_iters(15, 1, xlsx_file=excel_file, plot_file=plt_path)
    experiment1_iters(50, 1, xlsx_file=excel_file, plot_file=plt_path)
    experiment1_iters(100, 1, xlsx_file=excel_file, plot_file=plt_path)

    plt_path = f'{plot_folder}/ex22'
    experiment2_no_particles(10, 0, xlsx_file=excel_file, plot_file=plt_path)
    experiment2_no_particles(30, 0, xlsx_file=excel_file, plot_file=plt_path)
    experiment2_no_particles(100, 0, xlsx_file=excel_file, plot_file=plt_path)

    experiment2_no_particles(10, 1, xlsx_file=excel_file, plot_file=plt_path)
    experiment2_no_particles(30, 1, xlsx_file=excel_file, plot_file=plt_path)
    experiment2_no_particles(100, 1, xlsx_file=excel_file, plot_file=plt_path)

    plt_path = f'{plot_folder}/ex33'
    experiment3_inertia(0.1, 0, xlsx_file=excel_file, plot_file=plt_path)
    experiment3_inertia(0.5, 0, xlsx_file=excel_file, plot_file=plt_path)
    experiment3_inertia(0.9, 0, xlsx_file=excel_file, plot_file=plt_path)

    experiment3_inertia(0.1, 1, xlsx_file=excel_file, plot_file=plt_path)
    experiment3_inertia(0.5, 1, xlsx_file=excel_file, plot_file=plt_path)
    experiment3_inertia(0.9, 1, xlsx_file=excel_file, plot_file=plt_path)

    plt_path = f'{plot_folder}/ex44'
    experiment4_cognitive(0.1, 0, xlsx_file=excel_file, plot_file=plt_path)
    experiment4_cognitive(0.5, 0, xlsx_file=excel_file, plot_file=plt_path)
    experiment4_cognitive(0.9, 0, xlsx_file=excel_file, plot_file=plt_path)
    experiment4_cognitive(1.5, 0, xlsx_file=excel_file, plot_file=plt_path)
    experiment4_cognitive(2, 0, xlsx_file=excel_file, plot_file=plt_path)

    experiment4_cognitive(0.1, 1, xlsx_file=excel_file, plot_file=plt_path)
    experiment4_cognitive(0.5, 1, xlsx_file=excel_file, plot_file=plt_path)
    experiment4_cognitive(0.9, 1, xlsx_file=excel_file, plot_file=plt_path)
    experiment4_cognitive(1.5, 1, xlsx_file=excel_file, plot_file=plt_path)
    experiment4_cognitive(2, 1, xlsx_file=excel_file, plot_file=plt_path)

    plt_path = f'{plot_folder}/ex55'
    experiment5_social(0.1, 0, xlsx_file=excel_file, plot_file=plt_path)
    experiment5_social(0.5, 0, xlsx_file=excel_file, plot_file=plt_path)
    experiment5_social(0.9, 0, xlsx_file=excel_file, plot_file=plt_path)
    experiment5_social(1.5, 0, xlsx_file=excel_file, plot_file=plt_path)
    experiment5_social(2, 0, xlsx_file=excel_file, plot_file=plt_path)

    experiment5_social(0.1, 1, xlsx_file=excel_file, plot_file=plt_path)
    experiment5_social(0.5, 1, xlsx_file=excel_file, plot_file=plt_path)
    experiment5_social(0.9, 1, xlsx_file=excel_file, plot_file=plt_path)
    experiment5_social(1.5, 1, xlsx_file=excel_file, plot_file=plt_path)
    experiment5_social(2, 1, xlsx_file=excel_file, plot_file=plt_path)


def run_all_experiments():
    experiment1_iters(15, 0)
    experiment1_iters(50, 0)
    experiment1_iters(100, 0)

    experiment1_iters(15, 1)
    experiment1_iters(50, 1)
    experiment1_iters(100, 1)

    experiment2_no_particles(10, 0)
    experiment2_no_particles(30, 0)
    experiment2_no_particles(100, 0)

    experiment2_no_particles(10, 1)
    experiment2_no_particles(30, 1)
    experiment2_no_particles(100, 1)

    experiment3_inertia(0.1, 0)
    experiment3_inertia(0.5, 0)
    experiment3_inertia(0.9, 0)

    experiment3_inertia(0.1, 1)
    experiment3_inertia(0.5, 1)
    experiment3_inertia(0.9, 1)

    experiment4_cognitive(0.1, 0)
    experiment4_cognitive(0.5, 0)
    experiment4_cognitive(0.9, 0)
    experiment4_cognitive(1.5, 0)
    experiment4_cognitive(2, 0)

    experiment4_cognitive(0.1, 1)
    experiment4_cognitive(0.5, 1)
    experiment4_cognitive(0.9, 1)
    experiment4_cognitive(1.5, 1)
    experiment4_cognitive(2, 1)

    experiment5_social(0.1, 0)
    experiment5_social(0.5, 0)
    experiment5_social(0.9, 0)
    experiment5_social(1.5, 0)
    experiment5_social(2, 0)

    experiment5_social(0.1, 1)
    experiment5_social(0.5, 1)
    experiment5_social(0.9, 1)
    experiment5_social(1.5, 1)
    experiment5_social(2, 1)


if __name__ == '__main__':
    run_all_experiments()
